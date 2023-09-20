# -*- coding: utf-8 -*-
"""
Created on Sat Jul 10 23:37:44 2021

@author: Fanyi Sun
"""

from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border


def generate_spreadsheet(total_df, excel_path):
    total_df.to_excel(excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # set the general format
    for row in ws:
        for cell in row:
            align = Alignment(horizontal='center', vertical='center',
                              wrap_text=True)
            cell.alignment = align
            
    colA = ws['A']
    ws.merge_cells('A1:B1')
    titles = ['GENERAL', 'WEIGHTS AND DIMENSIONS',
              'VEHICLE AND RIG CONFIGURATION', 'ALIGNMENT DATA',
              'TYRES AND RIMS', 'USER LABELS', 'SETUP 1 INFORMATION']
    
    for cell in colA:
        if cell.value:  # if not a NoneType
            if 'Test' in cell.value or cell.value in titles:
                align = Alignment(horizontal='left', vertical='top')
                cell.alignment = align
                font = Font(size=16, bold=True, color='DC143C')
                cell.font = font

                if cell.value in titles:
                    ws.merge_cells(start_row=cell.row,start_column=cell.column,
                                   end_row=cell.row,end_column=cell.column+1)
                    # make the right border invisible
                    cellB = ws.cell(row=cell.row, column=cell.column+1)
                    border = Border(right=None)
                    cellB.border = border
                    
            else:
                cellB = ws.cell(row=cell.row, column=cell.column+1)
                # if nothing here then merge them
                if not cellB.value:
                    ws.merge_cells(start_row=cell.row,start_column=cell.column,
                                   end_row=cell.row,end_column=cell.column+1)
         
    ws.column_dimensions['A'].width = 51
    
    ws.row_dimensions[1].height = 100

    font = Font(size=14, bold=True, color='000000')
    ws['A1'].font = font
    # align = Alignment(horizontal='center',vertical='center')
    # ws['A1'].alignment = align
    
    ws.auto_filter.ref = 'B1'
    wb.save(excel_path)
    
    
if __name__ == '__main__':
    excel_path = r'Vehicle Data.xlsx'
    # only work if total_df already in variable
    generate_spreadsheet(total_df, excel_path)